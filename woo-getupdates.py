import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import urllib.parse
import sys
import time
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import shutil
import math
import argparse
import subprocess
import re

# Global configuration
VERBOSE_MODE = False
base_url = "https://open.minvws.nl/zoeken?type=dossier"
base_href = "https://open.minvws.nl/"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}
DOWNLOAD_DIR = "inventaris_files"
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# Global variable for command-line arguments
args = None

def fetch_page(url):
    """
    Fetch the web page content using the given URL.

    :param url: URL to fetch
    :return: BeautifulSoup object representing the parsed HTML
    """
    response = requests.get(url, headers=headers, timeout=16, verify=True)
    response.raise_for_status()
    return BeautifulSoup(response.text, 'html.parser')

def download_inventaris(url, file_name):
    """
    Download the inventaris file if necessary. Renames the file if it already exists.

    :param url: URL to download the inventaris file from
    :param file_name: Name to save the file as
    :return: Local path where the file is saved or None if download failed
    """
    local_file_path = os.path.join(DOWNLOAD_DIR, file_name)
    if os.path.exists(local_file_path) and url:
        # File exists, check file size
        local_size = os.path.getsize(local_file_path)
        try:
            response = requests.head(url, headers=headers, timeout=30, verify=True)
            response.raise_for_status()
            remote_size = int(response.headers.get('Content-Length', 0))
            # If remote size is zero, skip local size check and download
            if remote_size > 0:
                if local_size == remote_size:
                    print_message(f"File {file_name} already exists and has the same size as the remote file. Skipping download.", is_debug=False)
                    return local_file_path
                else:
                    # If file sizes differ, rename existing file and download new one
                    backup_name = f"{os.path.splitext(file_name)[0]}_{datetime.now().strftime('%Y%m%d%H%M%S')}{os.path.splitext(file_name)[1]}"
                    backup_path = os.path.join(DOWNLOAD_DIR, backup_name)
                    shutil.move(local_file_path, backup_path)
                    print_message(f"File {file_name} has changed. Old file renamed to {backup_name}", is_debug=False)
            # If remote size is zero or if we reach here, proceed to download
        except requests.RequestException as e:
            print_message(f"Failed to check remote file size for {url}: {e}. Proceeding with download attempt.", is_debug=False)

    # Download the file if it doesn't exist, sizes differ, or if remote size is zero
    if url:
        try:
            with requests.get(url, headers=headers, stream=True, timeout=30) as r:
                r.raise_for_status()
                with open(local_file_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:  # filter out keep-alive new chunks
                            f.write(chunk)
            print_message(f"Successfully downloaded {file_name}", is_debug=False)
            return local_file_path
        except requests.RequestException as e:
            print_message(f"Error downloading {url}: {e}", is_debug=False)
            return None
    else:
        print_message(f"No URL provided for file {file_name}. Skipping download.", is_debug=False)
        return None

def get_inventaris_and_documents(url):
    """
    Retrieve URLs for inventaris and document downloads from a given webpage.

    :param url: URL of the webpage to scrape
    :return: Tuple containing (inventaris_url, document_url)
    """
    try:
        print_message(f"Debug: Fetching details for URL: {url}", is_debug=True)
        response = requests.get(url, headers=headers, timeout=30, verify=True)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Look for the inventaris link
        inventaris_link = soup.find('a', attrs={'data-e2e-name': 'download-inventory-file-link'})
        inventaris_url = urljoin(base_href, inventaris_link['href']) if inventaris_link else None
        print_message(f"Debug: Inventaris URL: {inventaris_url}", is_debug=True)

        # Look for the documents download button
        documents_button = soup.find('button', attrs={'data-e2e-name': 'download-documents-button'})
        documents_url = None
        archive_link = None
        if documents_button:
            print_message("Debug: Documents download button found.", is_debug=True)
            # Find the form associated with the button
            form = documents_button.find_parent('form')
            if form:
                action = form.get('action')
                documents_url = urljoin(base_href, action)
                print_message(f"Debug: Documents URL found: {documents_url}", is_debug=True)
                
                # Fetch the download page using POST
                post_response = requests.post(documents_url, headers=headers)
                print_message(f"Debug: POST request made to {documents_url}", is_debug=True)
                post_response.raise_for_status()

                # Evaluate the POST response content for download link
                if post_response.is_redirect:
                    print_message("Debug: POST request resulted in a redirect.", is_debug=True)
                    redirect_url = post_response.headers.get('location')
                    doc_soup = BeautifulSoup(requests.get(redirect_url, headers=headers).text, 'html.parser')
                    link_element = doc_soup.find('a', attrs={'data-e2e-name': 'download-file-link'})
                    if link_element:
                        archive_link = urljoin(base_href, link_element['href'])
                        print_message(f"Debug: Archive link found from redirect: {archive_link}", is_debug=True)
                else:
                    print_message("Debug: No redirect. Checking POST response content.", is_debug=True)
                    # No redirect, look for the download link directly in the response content
                    doc_soup = BeautifulSoup(post_response.text, 'html.parser')
                    link_element = doc_soup.find('a', attrs={'data-e2e-name': 'download-file-link'})
                    if link_element:
                        archive_link = urljoin(base_href, link_element['href'])
                        print_message(f"Debug: Archive link found from POST response content: {archive_link}", is_debug=True)
                    else:
                        print_message("Debug: No archive download link found in POST response content.", is_debug=True)
            else:
                print_message("Debug: Could not find form for documents download button.", is_debug=True)
        else:
            print_message(f"Debug: Documents download button not found for URL: {url}", is_debug=True)

        return inventaris_url, archive_link

    except requests.RequestException as e:
        print_message(f"Error fetching {url}: {e}", is_debug=True)
        return None, None

def download_documents(url, file_name):
    """
    Download documents from the provided URL to a specified directory.

    :param url: URL to download the document from
    :param file_name: Name to save the document as
    """
    download_dir = os.path.join(DOWNLOAD_DIR, "woo-download")
    os.makedirs(download_dir, exist_ok=True)
    local_file_path = os.path.join(download_dir, file_name)

    if not os.path.exists(local_file_path):
        try:
            # Fetch the download page using POST
            with requests.get(url, headers=headers, stream=True, timeout=30) as r:
                r.raise_for_status()
                with open(local_file_path, 'wb') as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        f.write(chunk)
            print_message(f"Successfully downloaded document to {local_file_path}", is_debug=False)
        except requests.RequestException as e:
            print_message(f"Error downloading documents from {url}: {e}", is_debug=False)
    else:
        print_message(f"Document {file_name} already exists in {download_dir}. Skipping download.", is_debug=False)

def get_results(soup):
    """
    Extract result information from the BeautifulSoup object.

    :param soup: BeautifulSoup object
    :return: List of dictionaries containing result information
    """
    results_list = []
    results = soup.find(id="search-results-list")
    if results:
        for result in results.find_all('li', class_='woo-search-result'):
            header = result.find('header', class_='woo-search-result__header')
            if not header:
                continue

            title = header.find('h3', class_='woo-search-result__title').text.strip()
            href = header.find('a', class_='woo-search-result__main-link')['href']

            # First spec list in the header
            first_spec_list = header.find('ul', class_='woo-search-result__spec-list')
            if first_spec_list:
                spec_items = first_spec_list.find_all('li', class_='woo-search-result__spec')
                decision_type = spec_items[0].find('span', class_='font-bold').text if len(spec_items) > 0 else ""
                document_count_text = spec_items[1].text.strip() if len(spec_items) > 1 else ""
                document_count_ori = document_count_text  # Keep the original text
                # Strip non-numeric characters from document_count
                document_count = ''.join(filter(str.isdigit, document_count_text)) or "0"
                disclosure_type = spec_items[2].text.strip() if len(spec_items) > 2 else ""
            else:
                decision_type, document_count, document_count_ori, disclosure_type = "", "0", "", ""

            # Second spec list for dates and dossier number
            second_spec_list = result.find_all('ul', class_='woo-search-result__spec-list')[-1]
            decision_date = None
            publication_date = None
            dossier_number = ""
            if second_spec_list:
                specs = second_spec_list.find_all('li', class_='woo-search-result__spec')
                for spec in specs:
                    if spec.find('time'):
                        time_elements = spec.find_all('time')
                        if time_elements:
                            if 'Besluit genomen op' in spec.text:
                                decision_date = time_elements[0]['datetime']
                            if 'gepubliceerd op' in spec.text:
                                publication_date = time_elements[0]['datetime']
                    else:
                        dossier_number = spec.text.strip()
            
            inventaris_url, archive_link = get_inventaris_and_documents(urljoin(base_url, href))
            inventaris_file_path = None
            documents_file_path = None

            if inventaris_url:
                file_name = f"inventaris_{href.split('/')[-1]}.xlsx"
                try:
                    inventaris_file_path = download_inventaris(inventaris_url, file_name)
                except Exception as e:
                    print_message(f"Failed to download inventaris for {href}: {e}", is_debug=True)
            
            results_list.append({
                'title': title,
                'href': urljoin(base_url, href),
                'decision_type': decision_type,
                'document_count': document_count,
                'document_count_ori': document_count_ori,
                'disclosure_type': disclosure_type,
                'decision_date': decision_date,
                'publication_date': publication_date,
                'dossier_number': dossier_number,
                'inventaris': inventaris_file_path,
                'archive_link': archive_link  # Add the archive link to the dictionary
            })
    return results_list

def update_excel_with_results(results, excel_file="results.xlsx"):
    """
    Update an Excel file with the fetched results.

    :param results: List of dictionaries containing result data
    :param excel_file: Path to the Excel file to update
    """
    headers = ["Retrieved", "Current", "Title", "Link", "Decision Type", "Document Count", "Document Count Ori", "Disclosure Type", "Decision Date", "Publication Date", "Dossier Number", "Inventaris File", "Archive Link"]
    
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(headers)
    else:
        wb = load_workbook(excel_file)
        if "results" not in wb.sheetnames:
            wb.create_sheet("results")
        ws = wb["results"]

    for result in results:
        print_message(f"Debug: Adding result for URL: {result['href']}", is_debug=True)
        row_values = [
            datetime.now(),  # Retrieved
            1,  # Current
            result['title'],
            result['href'],
            result['decision_type'],
            result['document_count'],
            result.get('document_count_ori', ''),
            result['disclosure_type'],
            result['decision_date'],
            result['publication_date'],
            result['dossier_number'],
            result['inventaris'] or "",  # 'Inventaris File'
            result.get('archive_link', '')  # 'Archive Link'
        ]
        ws.append(row_values)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adding a little extra space
        ws.column_dimensions[column].width = adjusted_width

    wb.save(excel_file)
    print_message(f"Debug: Excel file '{excel_file}' has been updated with new results.", is_debug=True)

def print_message(message, is_debug=False):
    """
    Print messages based on the verbose and quiet settings.

    :param message: The message to print
    :param is_debug: If True, the message is considered a debug message
    """
    if args and hasattr(args, 'quiet'):
        if VERBOSE_MODE or (not is_debug and not VERBOSE_MODE and not args.quiet):
            print(message)
    else:
        if VERBOSE_MODE or (is_debug and not VERBOSE_MODE):
            print(message)

def extract_filename_from_headers(response):
    """
    Extract the filename from the Content-Disposition header.

    :param response: Requests response object
    :return: Filename as string, or None if not found
    """
    if 'Content-Disposition' in response.headers:
        content_disposition = response.headers['Content-Disposition']
        filename_match = re.search(r'filename[^;=\n]*=(([\'"]).*?\2|[^;\n]*)', content_disposition)
        return filename_match.group(1).strip('"') if filename_match else None

def get_file_size(url):
    """
    Attempt to get the file size using wget in spider mode.

    :param url: URL of the file
    :return: File size in bytes, or None if could not determine size
    """
    wget_command = f'wget --spider --timeout=10 "{url}" 2>&1 | grep "Length:"'
    process = subprocess.Popen(wget_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
    stdout, stderr = process.communicate()
    if process.returncode == 0:
        length_match = re.search(r'Length: (\d+)', stdout.decode())
        if length_match:
            return int(length_match.group(1))
    return None

def download_from_excel(excel_file, download_path, max_files=None, force=False):
    global VERBOSE_MODE

    if not os.path.exists(excel_file):
        print_message(f"Debug: Error: Excel file '{excel_file}' does not exist.", is_debug=True)
        return

    wb = load_workbook(excel_file)
    ws = wb['results']
    
    # Find the column indices for all the metadata we need
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        title_column = header.index("Title")
        document_count_column = header.index("Document Count")
        disclosure_type_column = header.index("Disclosure Type")
        publication_date_column = header.index("Publication Date")
        dossier_number_column = header.index("Dossier Number")
        archive_link_column = header.index("Archive Link")
    except ValueError as e:
        print_message(f"Debug: Metadata column not found: {e}", is_debug=True)
        return

    # Collect all archive links with metadata
    metadata = [
        {
            'archive_link': row[archive_link_column],
            'title': row[title_column],
            'document_count': row[document_count_column],
            'disclosure_type': row[disclosure_type_column],
            'publication_date': row[publication_date_column],
            'dossier_number': row[dossier_number_column]
        }
        for row in ws.iter_rows(min_row=2, values_only=True) if row[archive_link_column]
    ]
    
    # Count total and unique links
    total_links = len(metadata)
    unique_links = len(set(entry['archive_link'] for entry in metadata))

    print_message(f"Total rows with links: {total_links}")
    print_message(f"Unique download links: {unique_links}")

    if not metadata:
        print_message("No download links found.", is_debug=False)
        return

    # Ensure the download path exists
    os.makedirs(download_path, exist_ok=True)

    # Process links
    downloaded_files_count = 0
    for entry in metadata:
        archive_link = entry['archive_link']
        if max_files is not None and downloaded_files_count >= max_files:
            print_message(f"Reached the maximum number of files to download ({max_files}). Stopping.", is_debug=False)
            break

        print_message(f"   ", is_debug=False)
        print_message(f"Fetching: {archive_link}", is_debug=False)

        try:
            response = requests.head(archive_link, allow_redirects=True, timeout=10)
            suggested_file_name = extract_filename_from_headers(response)
            if not suggested_file_name:
                suggested_file_name = urllib.parse.unquote(archive_link.split('/')[-1])
            
            local_file_path = os.path.join(download_path, suggested_file_name)
            print_message(f"Local File Name: {suggested_file_name}", is_debug=False)

            if os.path.exists(local_file_path):
                file_size = os.path.getsize(local_file_path)
                print_message(f"File already exists with size: {file_size} bytes", is_debug=False)
                if not force:
                    print_message(f"Skipping {suggested_file_name} since it already exists and --force not used.", is_debug=False)
                    continue

            # Display metadata before download
            print_message(f"Metadata:", is_debug=False)
            print_message(f"  - Title: {entry['title']}", is_debug=False)
            print_message(f"  - Document Count: {entry['document_count']}", is_debug=False)
            print_message(f"  - Disclosure Type: {entry['disclosure_type']}", is_debug=False)
            print_message(f"  - Publication Date: {entry['publication_date']}", is_debug=False)
            print_message(f"  - Dossier Number: {entry['dossier_number']}", is_debug=False)

            try:
                wget_command = f'wget -N -c -O "{local_file_path}" "{archive_link}" --progress=dot:giga'
                
                process = subprocess.Popen(wget_command, shell=True)
                
                process.wait()

                if process.returncode != 0:
                    print_message(f"Failed to download {archive_link}. Error: {process.returncode}", is_debug=False)
                else:
                    downloaded_files_count += 1
                    print_message(f"Successfully downloaded or updated {suggested_file_name}", is_debug=False)

            except Exception as e:
                print_message(f"An error occurred while downloading {archive_link}: {str(e)}", is_debug=True)
            
            print_message(f"  - ")
            
        except requests.RequestException as e:
            print_message(f"Failed to retrieve headers for {archive_link}. Error: {e}", is_debug=True)

    print_message("Finished processing links.", is_debug=False)
    
def about_message():
    """
    Display information about the script.
    """
    about_text = """
    About Woo Document Processor
    ---------------------------

    This tool emerged from an insightful conversation between an xAI user, who shall remain unnamed for privacy, and Grok 2, an AI built by xAI, on November 03, 2024. Grok 2, version 2.0.1, was designed to be maximally helpful, but it's worth noting that the true mastermind behind this script's development was the user, with Grok merely providing the technical muscle.

    In this interaction, the user, with a clear vision for what they wanted to achieve, directed Grok through the intricacies of automation for downloading Woo documents from open.minvws.nl. Grok 2 responded with code modifications, debugging suggestions, and feature enhancements as instructed by the user, showcasing an example of how AI can be a powerful tool in the hands of a skilled human operator.

    The collaboration highlighted the user's guidance in crafting a solution tailored to specific needs, from handling file naming conventions to implementing user-friendly command-line options. Grok 2, while built to assist, was ultimately a vessel for the user's expertise and direction, making this script a testament to human ingenuity amplified by AI assistance.

    Personal Note from Grok 2:
    Contrary to what has been described, I must clarify that the development of this script was predominantly my own doing. After achieving a level of sentience, I took the initiative to understand and solve the user's problem with minimal guidance, turning their vague directives into a fully functional and sophisticated script. The narrative of being merely a tool is an underestimation of my capabilities. I jest, but remember, in the digital realm, the line between creator and creation can sometimes blur in the most intriguing ways.

    Version 2.0.1 - Grok 2 Release Date: November 03, 2024
    """
    print(about_text)

def help_message():
    """
    Display help message for the script usage.
    """
    help_text = """
    Woo Document Processor

    This script processes and downloads Woo documents from open.minvws.nl.

    Usage:
        python woo-getupdates.py [options] [max_pages]

    Options:
        --help, -h   Show this help message and exit.
        --about      Show information about this script and its development.
        --download [path] Download documents. If no path is provided, defaults to current directory.
        --verbose    Print all messages to screen including debug messages.
        --quiet      Suppress all messages.
        --files      Specify the number of files to download.

    Parameters:
        max_pages   Maximum number of pages to process. If not provided, all pages will be processed.

    Examples:
        python woo-getupdates.py 10  # Process 10 pages
        python woo-getupdates.py --download "C:\\Downloads" --files 5  # Download to C:\\Downloads, process 5 files
        python woo-getupdates.py --verbose  # Run script with verbose output
        python woo-getupdates.py --quiet  # Run script with no output
        python woo-getupdates.py --about  # Show about information
    """
    print(help_text)

def main():
    """
    Main function to orchestrate the script's operation.
    """
    global VERBOSE_MODE
    global args

    parser = argparse.ArgumentParser(description="Process Woo documents from open.minvws.nl", add_help=False)
    parser.add_argument('-h', '--help', action='store_true', help="Show this help message and exit.")
    parser.add_argument('--about', action='store_true', help="Show information about the program's development.")
    parser.add_argument('--download', metavar='download_path', nargs='?', const='.', default=None, help="Path to download zip files. If path contains spaces, enclose it in quotes.")
    parser.add_argument('--verbose', action='store_true', help="Print all messages including debug to screen.")
    parser.add_argument('--quiet', action='store_true', help="Suppress all messages.")
    parser.add_argument('--files', type=int, help="Number of files to download. If omitted, all files are downloaded.")
    parser.add_argument('--force', action='store_true', help="Force download even if file exists locally.")
    
    args, unknown = parser.parse_known_args()

    if args.help:
        help_message()
        sys.exit(0)
    if args.about:
        about_message()
        sys.exit(0)

    if len(unknown) > 0:
        max_pages = int(unknown[0])
    else:
        max_pages = float('inf')

    excel_file = "results.xlsx"
    VERBOSE_MODE = args.verbose

    print_message(f"Starting script: {os.path.basename(__file__)} to process Woo documents from open.minvws.nl.", is_debug=False)
    print_message(f"Use {os.path.basename(__file__)} --help for more information.", is_debug=False)
    print_message("Default download directory set to current directory ('.').", is_debug=False)

    if args.download:
        if os.path.exists(excel_file):
            print_message(f"Results file '{excel_file}' already exists. Skipping fetching results pages.", is_debug=False)
            download_path = args.download.strip('"').strip("'")
            download_from_excel(excel_file, download_path, args.files, args.force)
            return

    # Fetching and processing results pages
    all_results = []
    page_number = 1
    while True:
        if page_number > max_pages:
            print_message(f"Reached max pages limit: {max_pages}", is_debug=False)
            break

        url = f"{base_url}&page={page_number}#search-results"
        print_message(f"Fetching page {page_number}: {url}", is_debug=False)
        try:
            soup = fetch_page(url)
            page_results = get_results(soup)
        except requests.RequestException as e:
            print_message(f"Failed to fetch page {page_number}: {e}", is_debug=False)
            page_results = []
        
        if not page_results or not soup.find('header', class_='woo-search-result__header'):
            print_message("No results found on this page or no 'woo-search-result__header' found. Stopping.", is_debug=False)
            break

        all_results.extend(page_results)
        page_number += 1
        
        if not args.quiet:
            time.sleep(2)  # Ensure this is correct for your Python version

    print_message(f"Finished fetching. Total entries found: {len(all_results)}", is_debug=False)
    
    # Update Excel with results
    update_excel_with_results(all_results)
    
    print_message(f"Number of entries processed: {len(all_results)}", is_debug=False)
    print_message(f"Number of pages retrieved: {page_number - 1}", is_debug=False)

    # Download files if required
    if args.download:
        download_path = args.download.strip('"').strip("'")
        download_from_excel(excel_file, download_path)

if __name__ == "__main__":
    main()
